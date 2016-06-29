#!/usr/bin/python

"""
name: pan_policy_audit
description: Connect to a firewall/Panorama and export the number of rules using App-ID, User-ID, and threat profiles
author: Steve Barber
company: Palo Alto Networks
prerequisites: python 2.7.5+ with the libraries listed in the 'import' statements.
last updated: 06/08/2016

Change log:
** 1.0 - created script

"""
try:
    import sys
    import os.path
    import getpass
    import httplib
    import argparse
    import urllib
    import urllib2
    from lxml import etree
    import xml.etree.ElementTree as ET
    import re
    import logging
    import ssl
    from collections import defaultdict, OrderedDict
    import datetime
    import openpyxl
    from openpyxl.chart import BarChart, Series, Reference
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    from openpyxl.cell import get_column_letter
    from xml.dom import minidom

    ssl._create_default_https_context = ssl._create_unverified_context
except ImportError:
    raise ImportError("The following modules must be installed: lxml, openpyxl.  "
                      "Install them by running 'sudo pip install <module name>'")


def make_parser():
    """Construct the command line parser and return dictionary of args"""

    parser = argparse.ArgumentParser(
        description="Script to connect to a PAN-OS firewall and count rules using app-id, user-id, and threat profiles")
    parser.add_argument('-i', "--ip", help="IP address of Firewall")
    #parser.add_argument('-F', "--firewall", action="store_true", help="Firewall mode (default)")
    #parser.add_argument('-P', "--panorama", action="store_true", help="Panorama mode")
    parser.add_argument('-d', "--devicegroup", help="target specific Panorama device group")
    parser.add_argument('-u', "--user", help="username to login to Firewall; default 'admin'")
    parser.add_argument('-p', "--password", help="password for user; default 'admin'")
    parser.add_argument('-o', "--output", default='pan-policy-audit.xlsx', type=str,
                        help="Output data to .xlsx using the given file name")
    parser.add_argument('-T', "--threshold", default=0,
                        help="Flag status as 'In Progress' if percent is under given value")
    arguments = parser.parse_args()

    if arguments.ip:
        ip = arguments.ip
    else:
        #if arguments.panorama:
        #    arguments.panorama = raw_input("Enter the name or IP of Panorama: ")
        #else:
        #    print " "
        #    print "No arguments given. Defaulting to Firewall mode..."
        #    print " "
        #    arguments.firewall = raw_input("Enter the name or IP of the firewall: ")
        arguments.ip = raw_input("Enter the name or IP of the firewall or Panorama")
    if arguments.devicegroup:
        dg_specific = arguments.devicegroup
    if arguments.user:
        user = arguments.user
    else:
        arguments.user = raw_input("Enter the user login: ")
    if arguments.password:
        password = arguments.password
    else:
        arguments.password = getpass.getpass()

    arguments = vars(arguments)  # Convert namespace to dictionary
    return arguments


def send_api_request(url, values):
    # Function to send the api request to the firewall and return the
    # parsed response.
    data = urllib.urlencode(values)
    request = urllib2.Request(url, data, )
    response = urllib2.urlopen(request).read()
    return minidom.parseString(response)


def key_grab(ip, user, password):
    """Take in input ip, uname and password and retrieve API key from firewall"""

    url = 'https://' + ip + '/api'
    values = {'type': 'keygen', 'user': user, 'password': password}
    parsedKey = send_api_request(url, values)
    nodes = parsedKey.getElementsByTagName('key')
    api_key = nodes[0].firstChild.nodeValue

    return api_key


def get_sys_info(ip, key):

    url = 'https://' + ip + '/api'
    values = {'type': 'op', 'cmd': '<show><system><info></info></system></show>', 'key': key}
    parsedKey = send_api_request(url, values)
    hostname = parsedKey.getElementsByTagName('hostname')
    serial = parsedKey.getElementsByTagName('serial')
    model = parsedKey.getElementsByTagName('model')
    version = parsedKey.getElementsByTagName('sw-version')

    try:
        if model[0].firstChild.nodeValue not in ('Panorama', 'M-100', 'M-500'):
            mode = "f"
            print "Getting firewall sysinfo..."
        else:
            mode = "p"
            print "Getting Panorama sysinfo..."

        print " "
        print "System Info"
        print "---------------------------------------"
        print "Hostname: " + hostname[0].firstChild.nodeValue
        print "Serial: " + serial[0].firstChild.nodeValue
        print "Model: " + model[0].firstChild.nodeValue
        print "Software: " + version[0].firstChild.nodeValue
        print "---------------------------------------"
        print " "

        return hostname[0].firstChild.nodeValue, mode
    except:
        print("Unable to get system information.  Check credentials and try again")


def get_dg(ip, key, dg_specific):

    conn = httplib.HTTPSConnection(ip, context=ssl._create_unverified_context())
    print dg_specific
    if not dg_specific:
        """Panorama: Get device-groups."""
        request_str = "/api/?type=op&cmd=<show><devicegroups><%2Fdevicegroups><%2Fshow>&key="
        conn.request("GET", request_str + key)
        r = conn.getresponse()
        data = r.read()
        p_dg = etree.fromstring(data)
        conn.close()
        r.close()

    else:
        """Panorama: Get specific device-group."""
        dg_specific_encode = dg_specific.replace(" ", "%20")
        request_str = "/api/?type=op&cmd=<show><devicegroups><name>" + dg_specific_encode + "<%2Fname><%2Fdevicegroups><%2Fshow>&key="
        conn.request("GET", request_str + key)
        r = conn.getresponse()
        data = r.read()
        p_dg = etree.fromstring(data)
        conn.close()
        r.close()

    for x in p_dg.findall("result/devicegroups/entry"):
        dgname_encode = x.attrib['name'].replace(" ", "%20")
        dgname = x.attrib['name']
        dg_dict[dgname] = {}
        rule_dict[dgname] = {}
        for y in x.findall("devices/entry"):
            dg_dict[dgname][y.attrib['name']] = {}
            dg_dict[dgname][y.attrib['name']]['name_encoded'] = x.attrib['name'].replace(" ", "%20")
            dg_dict[dgname]['decrypt'] = {}
            rule_dict[dgname]['dcrypt_rule_count'] = 0
            rule_dict[dgname]['no_dcrypt_rule_count'] = 0
            rule_dict[dgname]['sec_rule_count'] = 0
            rule_dict[dgname]['features'] = {'virus': 0, 'spyware': 0, 'vuln': 0, 'url-filter': 0,
                                                       'data-filter': 0, 'file-block': 0, 'wildfire': 0, 'app-id': 0,
                                                       'profiles': 0, 'user-id': 0}

            for z in y:
                if z.tag == "serial":
                    dg_dict[dgname][y.attrib['name']][z.tag] = z.text
                if z.tag == "connected":
                    dg_dict[dgname][y.attrib['name']][z.tag] = z.text
                if z.tag == "hostname":
                    dg_dict[dgname][y.attrib['name']][z.tag] = z.text
                if z.tag == "ip-address":
                    dg_dict[dgname][y.attrib['name']][z.tag] = z.text
                if z.tag == "model":
                    dg_dict[dgname][y.attrib['name']][z.tag] = z.text
                if z.tag == "sw-version":
                    dg_dict[dgname][y.attrib['name']][z.tag] = z.text
    #for x in dg_dict:
    #    for y in dg_dict[x].items():
    #        print x
    #        print '-->' + str(y)
    #        print ""

    for dg in dg_dict:
        for device in dg_dict[dg]:
            if dg_dict[dg][device].get('connected') == 'yes':
                if dg not in active_dg:
                    active_dg.append(dg)

    for dg in active_dg:
        parse_dg(ip, key, dg)


def parse_dg(ip, key, dg):

    conn = httplib.HTTPSConnection(ip, context=ssl._create_unverified_context())

    print ""
    print "Grabbing copy of shared pre-rulebase..."
    conn.request("GET", "/api/?type=config&action=get&xpath=/config/shared/pre-rulebase&key=" + key)
    r1 = conn.getresponse()
    data1 = r1.read()
    p_shared_pre_rulebase = etree.fromstring(data1)
    conn.close()
    r1.close()

    print "Grabbing copy of shared post-rulebase..."
    print ""
    conn.request("GET", "/api/?type=config&action=get&xpath=/config/shared/post-rulebase&key=" + key)
    r2 = conn.getresponse()
    data2 = r2.read()
    p_shared_post_rulebase = etree.fromstring(data2)
    conn.close()
    r2.close()

    dgname_encode = dg.replace(" ", "%20")
    print "Grabbing copy of pre-rulebase for device-group '" + dg + "'."
    conn.request("GET",
                 "/api/?type=config&action=get&xpath=/config/devices/entry[@name='localhost.localdomain']"
                 "/device-group/entry[@name='" + dgname_encode + "']/pre-rulebase" + "&key=" + key)
    r1 = conn.getresponse()
    data1 = r1.read()
    dg_rulebase_dict[dg + "-pre"] = etree.fromstring(data1)
    conn.close()
    r1.close()

    print "Grabbing copy of post-rulebase for device-group '" + dg + "'."
    conn.request("GET",
                 "/api/?type=config&action=get&xpath=/config/devices/entry[@name='localhost.localdomain']"
                 "/device-group/entry[@name='" + dgname_encode + "']/post-rulebase" + "&key=" + key)
    r2 = conn.getresponse()
    data2 = r2.read()
    dg_rulebase_dict[dg + "-post"] = etree.fromstring(data2)
    conn.close()
    r2.close()

    print "Counting '" + dg + "' decrypt pre-rules"
    for rule in dg_rulebase_dict.get(dg + "-pre").findall('result/pre-rulebase/decryption/rules/'):
        if not rule.findall('disabled'):
            for profile in rule.findall('action'):
                if profile.text == "decrypt":
                    dg_dict[dg]['decrypt'] = 'yes'
                    rule_dict[dg]['dcrypt_rule_count'] += 1
                else:
                    rule_dict[dg]['no_dcrypt_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    for profile in rule.findall('action'):
                        if profile.text == "decrypt":
                            dg_dict[dg]['decrypt'] = 'yes'
                            rule_dict[dg]['dcrypt_rule_count'] += 1
                        else:
                            rule_dict[dg]['no_dcrypt_rule_count'] += 1

    print "Counting '" + dg + "' decrypt post-rules"
    for rule in dg_rulebase_dict.get(dg + "-post").findall('result/post-rulebase/decryption/rules/'):
        if not rule.findall('disabled'):
            for profile in rule.findall('action'):
                if profile.text == "decrypt":
                    dg_dict[dg]['decrypt'] = 'yes'
                    rule_dict[dg]['dcrypt_rule_count'] += 1
                else:
                    rule_dict[dg]['no_dcrypt_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    for profile in rule.findall('action'):
                        if profile.text == "decrypt":
                            dg_dict[dg]['decrypt'] = 'yes'
                            rule_dict[dg]['dcrypt_rule_count'] += 1
                        else:
                            rule_dict[dg]['no_dcrypt_rule_count'] += 1

    print "Counting shared decrypt pre-rules"
    for rule in p_shared_pre_rulebase.findall('result/pre-rulebase/decryption/rules/'):
        if not rule.findall('disabled'):
            for profile in rule.findall('action'):
                if profile.text == "decrypt":
                    dg_dict[dg]['decrypt'] = 'yes'
                    rule_dict[dg]['dcrypt_rule_count'] += 1
                else:
                    rule_dict[dg]['no_dcrypt_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    for profile in rule.findall('action'):
                        if profile.text == "decrypt":
                            dg_dict[dg]['decrypt'] = 'yes'
                            rule_dict[dg]['dcrypt_rule_count'] += 1
                        else:
                            rule_dict[dg]['no_dcrypt_rule_count'] += 1

    print "Counting shared decrypt post-rules"
    for rule in p_shared_post_rulebase.findall('result/post-rulebase/decryption/rules/'):
        if not rule.findall('disabled'):
            for profile in rule.findall('action'):
                if profile.text == "decrypt":
                    dg_dict[dg]['decrypt'] = 'yes'
                    rule_dict[dg]['dcrypt_rule_count'] += 1
                else:
                    rule_dict[dg]['no_dcrypt_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    for profile in rule.findall('action'):
                        if profile.text == "decrypt":
                            dg_dict[dg]['decrypt'] = 'yes'
                            rule_dict[dg]['dcrypt_rule_count'] += 1
                        else:
                            rule_dict[dg]['no_dcrypt_rule_count'] += 1

    print "Counting '" + dg + "' security pre-rules"
    for rule in dg_rulebase_dict.get(dg + "-pre").findall('result/pre-rulebase/security/rules/'):
        if not rule.findall('disabled'):
            rule_dict[dg]['sec_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    rule_dict[dg]['sec_rule_count'] += 1

    print "Counting '" + dg + "' security post-rules"
    for rule in dg_rulebase_dict.get(dg + "-post").findall('result/post-rulebase/security/rules/'):
        if not rule.findall('disabled'):
            rule_dict[dg]['sec_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    rule_dict[dg]['sec_rule_count'] += 1

    print "Counting shared security pre-rules"
    for rule in p_shared_pre_rulebase.findall('result/pre-rulebase/security/rules/'):
        if not rule.findall('disabled'):
            rule_dict[dg]['sec_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    rule_dict[dg]['sec_rule_count'] += 1

    print "Counting shared security post-rules"
    print ""
    for rule in p_shared_post_rulebase.findall('result/post-rulebase/security/rules/'):
        if not rule.findall('disabled'):
            rule_dict[dg]['sec_rule_count'] += 1
        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    rule_dict[dg]['sec_rule_count'] += 1

    #for y in dg_dict[x].items():
    #    print x
    #    print '-->' + str(y)
    #    print ""

    xmlpath = "result/pre-rulebase/security/rules/"
    loop_rules(dg, xmlpath, p_shared_pre_rulebase)

    xmlpath = "result/post-rulebase/security/rules/"
    loop_rules(dg, xmlpath, p_shared_post_rulebase)

    xmlpath = "result/pre-rulebase/security/rules/"
    dg_pre_rulebase = dg_rulebase_dict.get(dg + "-pre")
    loop_rules(dg, xmlpath, dg_pre_rulebase)

    xmlpath = "result/post-rulebase/security/rules/"
    dg_post_rulebase = dg_rulebase_dict.get(dg + "-post")
    loop_rules(dg, xmlpath, dg_post_rulebase)

    print_output(dg)


def get_f_info(ip, key):

    conn = httplib.HTTPSConnection(ip, context=ssl._create_unverified_context())

    request_str = "/api/?type=op&cmd=<show><system><info></info></system></show>&key="

    conn.request("GET", request_str + key)
    r = conn.getresponse()
    data = r.read()
    fw_info = etree.fromstring(data)

    print "Getting Firewall Rulebase..."
    print ""
    request_str = "/api/?type=config&action=get&xpath=/config/devices/entry[@name='localhost.localdomain']/vsys/entry" \
                  "[@name='vsys1']/rulebase&key="

    conn.request("GET", request_str + key)
    r = conn.getresponse()
    data = r.read()
    fw_rulebase = etree.fromstring(data)
    conn.close()

    for x in fw_info.findall("result/system/"):
        if x.tag == 'serial':
            fw_dict[x.text] = {}
            fw_dict[x.text][x.tag] = x.text
            fw_dict[x.text]['decrypt'] = {}

    for x in fw_dict:
        for y in fw_info.findall("result/system/"):
            if y.tag == 'hostname':
                fw_dict[x][y.tag] = y.text
            if y.tag == 'ip-address':
                fw_dict[x][y.tag] = y.text
            if y.tag == 'model':
                fw_dict[x][y.tag] = y.text
            if y.tag == 'sw-version':
                fw_dict[x][y.tag] = y.text
        rule_dict[x] = {}
        rule_dict[x]['dcrypt_rule_count'] = 0
        rule_dict[x]['no_dcrypt_rule_count'] = 0
        rule_dict[x]['sec_rule_count'] = 0
        rule_dict[x]['features'] = {'virus': 0, 'spyware': 0, 'vuln': 0, 'url-filter': 0, 'data-filter': 0,
                                         'file-block': 0, 'wildfire': 0, 'app-id': 0, 'profiles': 0, 'user-id': 0}


    for x in fw_dict:
        print "Counting " + x + " decrypt rules"
        for rule in fw_rulebase.findall('result/rulebase/decryption/rules/'):
            if not rule.findall('disabled'):
                for profile in rule.findall('action'):
                    if profile.text == "decrypt":
                        fw_dict[x]['decrypt'] = 'yes'
                        rule_dict[x]['dcrypt_rule_count'] += 1
                    else:
                        rule_dict[x]['no_dcrypt_rule_count'] += 1
            else:
                for profile in rule.findall('disabled'):
                    if profile.text != 'yes':
                        for profile in rule.findall('action'):
                            if profile.text == "decrypt":
                                fw_dict[x]['decrypt'] = 'yes'
                                rule_dict[x]['dcrypt_rule_count'] += 1
                            else:
                                rule_dict[x]['no_dcrypt_rule_count'] += 1

        print "Counting " + x + " security rules"
        print ""
        for rule in fw_rulebase.findall('result/rulebase/security/rules/'):
            if not rule.findall('disabled'):
                rule_dict[x]['sec_rule_count'] += 1
            else:
                for profile in rule.findall('disabled'):
                    if profile.text != 'yes':
                        rule_dict[x]['sec_rule_count'] += 1

        xmlpath = "result/rulebase/security/rules/"
        loop_rules(x, xmlpath, fw_rulebase)

        print_output(x)


def loop_rules(dg, xmlpath, rulebase):
    """Recurse through rulset and find features used"""
    state = "disabled"
    rulepath = "profile-setting/"

    for rule in rulebase.findall(xmlpath):
        if not rule.findall('disabled'):
            for profile in rule.findall(rulepath):
                for entry in profile.getchildren():
                    if entry.tag == 'virus':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['virus'] += 1
                    if entry.tag == 'spyware':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['spyware'] += 1
                    if entry.tag == 'vulnerability':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['vuln'] += 1
                    if entry.tag == 'url-filtering':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['url-filter'] += 1
                    if entry.tag == 'data-filtering':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['data-filter'] += 1
                    if entry.tag == 'file-blocking':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['file-block'] += 1
                    if entry.tag == 'wildfire-analysis':
                        for x in entry.getchildren():
                            if x.tag == 'member':
                                rule_dict[dg]['features']['wildfire'] += 1
            for app in rule.findall('application/'):
                if app.text != "any":
                    rule_dict[dg]['features']['app-id'] += 1
                    break

            for user in rule.findall('source-user/'):
                if user.text != "any":
                    rule_dict[dg]['features']['user-id'] += 1
                    break

        else:
            for profile in rule.findall('disabled'):
                if profile.text != 'yes':
                    for profile in rule.findall(rulepath):
                        for entry in profile.getchildren():
                            if entry.tag == 'virus':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['virus'] += 1
                            if entry.tag == 'spyware':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['spyware'] += 1
                            if entry.tag == 'vulnerability':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['vuln'] += 1
                            if entry.tag == 'url-filtering':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['url-filter'] += 1
                            if entry.tag == 'data-filtering':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['data-filter'] += 1
                            if entry.tag == 'file-blocking':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['file-block'] += 1
                            if entry.tag == 'wildfire-analysis':
                                for x in entry.getchildren():
                                    if x.tag == 'member':
                                        rule_dict[dg]['features']['wildfire'] += 1
                    for app in rule.findall('application/'):
                        if app.text != "any":
                            rule_dict[dg]['features']['app-id'] += 1
                            break

                    for user in rule.findall('source-user/'):
                        if user.text != "any":
                            rule_dict[dg]['features']['user-id'] += 1
                            break


def print_output(dg):
    """Print data to screen"""
    try:
        print "Total # of decryption rules for '" + dg + "': " + str(rule_dict[dg]['no_dcrypt_rule_count'] +
                                                                   rule_dict[dg]['dcrypt_rule_count'])
        print "------------------------------------------"
        print "  no-decrypt rules ----- " + str(rule_dict[dg]['no_dcrypt_rule_count'])
        print "  decrypt rules -------- " + str(rule_dict[dg]['dcrypt_rule_count'])
        print "------------------------------------------"
        print " "
        print "Total # of security rules for '" + dg + "': " + str(rule_dict[dg]['sec_rule_count'])
        print "------------------------------------------"
        if not rule_dict[dg]['sec_rule_count'] == 0:
            print "  # using app-id ------- " + str(rule_dict[dg]['features']['app-id']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['app-id']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "  # using user-id ------ " + str(rule_dict[dg]['features']['user-id']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['user-id']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "  - profiles: "
            print "      - antivirus ------ " + str(rule_dict[dg]['features']['virus']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['virus']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "      - spyware -------- " + str(rule_dict[dg]['features']['spyware']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['spyware']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "      - vulnerability -- " + str(rule_dict[dg]['features']['vuln']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['vuln']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "      - url-filtering -- " + str(rule_dict[dg]['features']['url-filter']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['url-filter']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "      - file-blocking -- " + str(rule_dict[dg]['features']['file-block']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['file-block']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "      - wildfire ------- " + str(rule_dict[dg]['features']['wildfire']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['wildfire']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "      - data-filtering - " + str(rule_dict[dg]['features']['data-filter']) + \
                  " ({:.1f}%".format(rule_dict[dg]['features']['data-filter']/float(rule_dict[dg]['sec_rule_count']) *
                                      100) + ")"
            print "------------------------------------------"
        else:
            print "No security rules found.  Not reporting threat profile information."
        print " "
        print " "
    except:
        sys.exit(1)


def xl_writer(filename, threshold):
    """Write data to Excel"""

    feature_list = None
    rule_count = None

    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ft_implemented = Font(color='3E511F')
        fill_implemented = PatternFill(start_color='CFDFAF', end_color='90EE90', fill_type='solid')
        ft_not_implemented = Font(color='822627')
        fill_not_implemented = PatternFill(start_color='DFA8A9', end_color='DFA8A9', fill_type='solid')
        ft_in_progress = Font(color='822627')
        fill_in_progress = PatternFill(start_color='FFFF0B', end_color='FFFF0B', fill_type='solid')

        for x in wb.sheetnames:
            if x == 'Graph':
                delsheet = wb.get_sheet_by_name('Graph')
                wb.remove_sheet(delsheet)

        if mode == 'p':
            for dg in active_dg:
                row_append = ws.max_row + 1
                for device in dg_dict[dg]:
                    try:
                        for a, b in dg_dict[dg][device].items():
                            if a == 'connected' and b == 'yes':
                                for x in dg_dict[dg][device]:
                                    if x == "serial":
                                        ws.cell(row=row_append, column=1).value = dg_dict[dg][device].get(x)
                                    if x == "hostname":
                                        ws.cell(row=row_append, column=2).value = dg_dict[dg][device].get(x)
                                    if x == 'model':
                                        ws.cell(row=row_append, column=3).value = dg_dict[dg][device].get(x)
                                    ws.cell(row=row_append, column=4).value = hostname
                                    ws.cell(row=row_append, column=5).value = dg
                                    ws.cell(row=row_append, column=6).value = rule_dict[dg]['sec_rule_count']
                                    rule_count = rule_dict[dg]['sec_rule_count']
                                    feature_list = [rule_dict[dg]['features']['app-id'], rule_dict[dg]['features']['user-id'],
                                                    rule_dict[dg]['features']['virus'], rule_dict[dg]['features']['spyware'],
                                                    rule_dict[dg]['features']['vuln'], rule_dict[dg]['features']['url-filter'],
                                                    rule_dict[dg]['features']['file-block'], rule_dict[dg]['features']['wildfire'],
                                                    rule_dict[dg]['features']['data-filter']]
                                    i = 7
                                    for x in feature_list:
                                        if x > 0:
                                            if (x/float(rule_count) * 100) > int(threshold):
                                                ws.cell(row=row_append, column=i).value = 'Implemented (' + str(x) + " : " + \
                                                                                          "{:.1f}%".format(x/float(rule_count) * 100) + ")"
                                                ws.cell(row=row_append, column=i).fill = fill_implemented
                                                ws.cell(row=row_append, column=i).font = ft_implemented
                                            else:
                                                ws.cell(row=row_append, column=i).value = 'In Progress (' + str(x) + " : " + \
                                                                                          "{:.1f}%".format(x/float(rule_count) * 100) + ")"
                                                ws.cell(row=row_append, column=i).fill = fill_in_progress
                                                ws.cell(row=row_append, column=i).font = ft_in_progress
                                        else:
                                            ws.cell(row=row_append, column=i).value = 'Not Implemented'
                                            ws.cell(row=row_append, column=i).fill = fill_not_implemented
                                            ws.cell(row=row_append, column=i).font = ft_not_implemented
                                        i += 1

                                    if dg_dict[dg]['decrypt'] == 'yes':
                                        ws.cell(row=row_append, column=16).value = 'Implemented'
                                        ws.cell(row=row_append, column=16).fill = fill_implemented
                                        ws.cell(row=row_append, column=16).font = ft_implemented
                                    else:
                                        ws.cell(row=row_append, column=16).value = 'Not Implemented'
                                        ws.cell(row=row_append, column=16).fill = fill_not_implemented
                                        ws.cell(row=row_append, column=16).font = ft_not_implemented

                                    i = 17
                                    for x in feature_list:
                                        ws.cell(row=row_append, column=i).value = x
                                        i += 1
                                row_append += 1
                    except:
                        pass

        else:
            for serial in fw_dict:
                row_append = ws.max_row + 1
                for a, b in fw_dict[serial].items():
                    if a == "serial":
                        ws.cell(row=row_append, column=1).value = b
                    if a == "hostname":
                        ws.cell(row=row_append, column=2).value = b
                    if a == 'model':
                        ws.cell(row=row_append, column=3).value = b
                ws.cell(row=row_append, column=4).value = "N/A"
                ws.cell(row=row_append, column=5).value = "N/A"
                ws.cell(row=row_append, column=6).value = rule_dict[serial]['sec_rule_count']
                rule_count = rule_dict[serial]['sec_rule_count']
                feature_list = [rule_dict[serial]['features']['app-id'], rule_dict[serial]['features']['user-id'],
                                rule_dict[serial]['features']['virus'], rule_dict[serial]['features']['spyware'],
                                rule_dict[serial]['features']['vuln'], rule_dict[serial]['features']['url-filter'],
                                rule_dict[serial]['features']['file-block'], rule_dict[serial]['features']['wildfire'],
                                rule_dict[serial]['features']['data-filter']]
                i = 7
                for x in feature_list:
                    if x > 0:
                        if (x/float(rule_count) * 100) > int(threshold):
                            ws.cell(row=row_append, column=i).value = 'Implemented (' + str(x) + " : " + \
                                                                      "{:.1f}%".format(x/float(rule_count) * 100) + ")"
                            ws.cell(row=row_append, column=i).fill = fill_implemented
                            ws.cell(row=row_append, column=i).font = ft_implemented
                        else:
                            ws.cell(row=row_append, column=i).value = 'In Progress (' + str(x) + " : " + \
                                                                      "{:.1f}%".format(x/float(rule_count) * 100) + ")"
                            ws.cell(row=row_append, column=i).fill = fill_in_progress
                            ws.cell(row=row_append, column=i).font = ft_in_progress
                    else:
                        ws.cell(row=row_append, column=i).value = 'Not Implemented'
                        ws.cell(row=row_append, column=i).fill = fill_not_implemented
                        ws.cell(row=row_append, column=i).font = ft_not_implemented
                    i += 1

                if fw_dict[serial]['decrypt'] == 'yes':
                    ws.cell(row=row_append, column=16).value = 'Implemented'
                    ws.cell(row=row_append, column=16).fill = fill_implemented
                    ws.cell(row=row_append, column=16).font = ft_implemented
                else:
                    ws.cell(row=row_append, column=16).value = 'Not Implemented'
                    ws.cell(row=row_append, column=16).fill = fill_not_implemented
                    ws.cell(row=row_append, column=16).font = ft_not_implemented

                i = 17
                for x in feature_list:
                    ws.cell(row=row_append, column=i).value = x
                    i += 1
                row_append += 1

        dims = {}
        for row in ws.rows:
            for cell in row:
                try:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
                except:
                    pass
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

        wb.save(filename)
        print "Data written to file: " + str(os.getcwd()) + "/" + filename
        print " "
    else:
        xl_create(filename)


def xl_create(filename):
    """Create .xls file if it doesn't exist"""
    print "Excel file (" + filename + ") doesn't exist.  Creating .xls and trying again..."
    print " "

    device_header = ['Serial Number', 'Hostname', 'Model', 'Panorama', 'Device Group', 'Security Rule Total']
    feature_header = ['App-ID', 'User-ID', 'Antivirus', 'Anti-Spyware', 'Vulnerability Protection', 'URL Filtering',
                      'File Blocking', 'WildFire', 'Data Filtering', 'SSL Decryption']
    hidden_header = ['App-ID #', 'User-ID #', 'Antivirus #', 'Anti-Spyware #', 'Vulnerability Protection #',
                     'URL Filtering #', 'File-Blocking #', 'Wildfire #', 'Data Filtering #']

    device_font = Font(color='FFFFFF')
    device_fill = PatternFill(start_color='6D6D6D', end_color='6D6D6D', fill_type='solid')
    feature_font = Font(color='FFFFFF')
    feature_fill = PatternFill(start_color='2A4C93', end_color='2A4C93', fill_type='solid')
    hidden_font = Font(color='FFFFFF')
    hidden_fill = PatternFill(start_color='D9550C', end_color='D9550C', fill_type='solid')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.row_dimensions[1].height = 50

    i = 0
    for item in device_header:
        i += 1
        ws.cell(row=1, column=i).value = item
        ws.cell(row=1, column=i).font = device_font
        ws.cell(row=1, column=i).fill = device_fill
    i = 6
    for item in feature_header:
        i += 1
        ws.cell(row=1, column=i).value = item
        ws.cell(row=1, column=i).font = feature_font
        ws.cell(row=1, column=i).fill = feature_fill
    i = 16
    for item in hidden_header:
        i += 1
        ws.cell(row=1, column=i).value = item
        ws.cell(row=1, column=i).font = hidden_font
        ws.cell(row=1, column=i).fill = hidden_fill
        ws.column_dimensions[get_column_letter(i)].hidden = True

    wb.save(filename)
    xl_writer(filename, parser['threshold'])


def main():
    """Main function"""
    global mode #PANORAMA VS DIRECT TO FIREWALL
    global hostname #PANORAMA HOSTNAME
    global dg_dict #DICTIONARY CONTAINING ALL DEVICE-GROUPS
    global fw_dict #DICTIONARY CONTAINING ALL DEVICE-GROUPS
    global active_dg #LIST CONTAINING DEVICE-GROUPS WITH FWs IN CONNECTED STATE
    global p_shared_pre_rulebase #DICTIONARY CONTAINING SHARED PRE-RULEBASE
    global p_shared_post_rulebase #DICTIONARY CONTAINING SHARED POST-RULEBASE
    global dg_rulebase_dict #DICTIONARY CONTAINING DEVICE-GROUP PRE AND POST RULEBASE
    global rule_dict #DICTIONARY CONTAINING DEVICE-GROUP SECURITY & DECRYPTION RULE COUNT

    mode = ""
    hostname = None
    dg_dict = OrderedDict()
    fw_dict = {}
    active_dg = []
    dg_rulebase_dict = {}
    rule_dict = defaultdict(int)

    try:
        global parser
        parser = make_parser()
    except:
        sys.exit(1)

    apikey = key_grab(parser['ip'], parser['user'], parser['password'])

    hostname, mode = get_sys_info(parser['ip'], apikey)

    if mode == "f":
        print "Parsing firewall..."
        get_f_info(parser['ip'], apikey)
        xl_writer(parser['output'], parser['threshold'])
    else:
        print "Parsing Panorama..."
        get_dg(parser['ip'], apikey, parser['devicegroup'])
        xl_writer(parser['output'], parser['threshold'])


if __name__ == '__main__':
    main()
